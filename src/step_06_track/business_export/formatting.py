"""Formato openpyxl del Excel ejecutivo: paletas + reglas condicionales.

Aplica estilos consistentes a todas las hojas del workbook ya escrito por
los `build_*_df`. Encapsulado aqui para que cuando se ajusten colores o
reglas condicionales, no haya que tocar la logica de DataFrames.
"""
from __future__ import annotations

# Paletas (bg, fg) en hex. Usadas en hoja Inicio + Acciones.
_VERDICT_COLORS = {
    "GREEN":   ("C8E6C9", "1B5E20"),
    "GREEN-2": ("DCEDC8", "33691E"),
    "AMBER":   ("FFE0B2", "BF360C"),
    "RED":     ("FFCDD2", "B71C1C"),
}

_SCORE_COLORS = {
    "ALTO":  ("C8E6C9", "1B5E20"),
    "MEDIO": ("FFE0B2", "BF360C"),
    "BAJO":  ("FFCDD2", "B71C1C"),
}

_SEVERITY_COLORS = {
    "CRÍTICO": ("FFCDD2", "B71C1C"),
    "ATENCIÓN": ("FFE0B2", "BF360C"),
    "OK": ("C8E6C9", "1B5E20"),
}


def apply_formatting(workbook) -> None:
    """Aplica estilos consistentes a todas las hojas del Excel."""
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="0C2A4D")  # navy
    header_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill("solid", fgColor="EEF3FA")
    section_font = Font(bold=True, color="0C2A4D", size=12)
    block_fill = PatternFill("solid", fgColor="C9A961")  # gold
    block_font = Font(bold=True, color="0C2A4D", size=14)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if ws.max_row < 1:
            continue

        # Header (fila 1) en navy + bold blanco
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"

        # Auto-width (cap 60 para hojas con texto largo, 40 para tablas)
        cap = 80 if sheet_name in ("Inicio", "Acciones", "Glosario") else 45
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
                default=10,
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(cap, max(12, max_len + 2))

        # Wrap text en las hojas narrativas
        if sheet_name in ("Inicio", "Acciones", "Glosario"):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                ws.row_dimensions[row[0].row].height = None  # auto

        # Hojas de predicciones: formato condicional en error %
        if sheet_name.startswith("Predicciones_"):
            header = [c.value for c in ws[1]]
            err_col_name = "Error porcentual (%)"
            if err_col_name in header:
                col_idx = header.index(err_col_name) + 1
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator="greaterThan", formula=["20"],
                    fill=PatternFill("solid", fgColor="FCE4E4"),
                    font=Font(color="9B1C1C", bold=True)))
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator="between", formula=["10", "20"],
                    fill=PatternFill("solid", fgColor="FFF5D6")))
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator="lessThan", formula=["10"],
                    fill=PatternFill("solid", fgColor="E0F2EE")))

        # Hoja Resumen: pintar secciones ##  ##
        if sheet_name == "Resumen":
            for row in ws.iter_rows(min_row=2):
                first = row[0].value
                if isinstance(first, str) and first.startswith("##"):
                    for c in row:
                        c.fill = section_fill
                        c.font = section_font

        # Hoja Inicio: pintar bloques ◆◆◆ y celdas de veredicto/score
        if sheet_name == "Inicio":
            for row in ws.iter_rows(min_row=2):
                first = row[0].value
                third = row[2].value if len(row) >= 3 else None
                if isinstance(first, str) and first.startswith("◆◆◆"):
                    for c in row[:3]:
                        c.fill = block_fill
                        c.font = block_font
                        c.alignment = Alignment(horizontal="left", vertical="center")
                    ws.row_dimensions[row[0].row].height = 28
                # Pintar celda de veredicto/score
                if isinstance(third, str):
                    key = third.upper()
                    if key in _VERDICT_COLORS:
                        bg, fg = _VERDICT_COLORS[key]
                        for c in row[:3]:
                            c.fill = PatternFill("solid", fgColor=bg)
                            c.font = Font(bold=True, color=fg, size=12)
                    elif key in _SCORE_COLORS:
                        bg, fg = _SCORE_COLORS[key]
                        row[2].fill = PatternFill("solid", fgColor=bg)
                        row[2].font = Font(bold=True, color=fg, size=11)
                        row[2].alignment = Alignment(horizontal="center")

        # Hoja Acciones: pintar columna severidad
        if sheet_name == "Acciones":
            header = [c.value for c in ws[1]]
            if "Severidad" in header:
                sev_col = header.index("Severidad") + 1
                for row in ws.iter_rows(min_row=2):
                    cell = row[sev_col - 1]
                    val = (cell.value or "").upper()
                    for key, (bg, fg) in _SEVERITY_COLORS.items():
                        if key in val:
                            cell.fill = PatternFill("solid", fgColor=bg)
                            cell.font = Font(bold=True, color=fg)
                            cell.alignment = Alignment(horizontal="center")
                            break

        # Borde sutil en todas las celdas con datos (excepto bloques pintados)
        if sheet_name not in ("Inicio",):
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value not in (None, ""):
                        cell.border = thin_border
