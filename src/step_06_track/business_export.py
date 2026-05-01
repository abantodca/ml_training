"""Exporta el detalle de predicciones a Excel para analisis del negocio.

Excel ejecutivo con 8 hojas (orden de lectura):

  1. Inicio              Portada con veredicto + 3 KPIs en lenguaje natural
                         + indice de hojas. Una persona no-tecnica entiende
                         el resultado leyendo SOLO esta hoja.
  2. Acciones            Recomendaciones auto-generadas desde el analisis
                         (subgrupos problematicos, overfitting, etc.).
  3. Resumen             Metricas tecnicas globales + estado vs targets.
  4. Por_FORMATO         Agregado por FORMATO (n, error, ranking peor->mejor).
  5. Por_FUNDO           Idem por FUNDO.
  6. Predicciones_OOF    Detalle fila a fila con OOF (honesto). Hoja
                         operacional para filtros + pivots del negocio.
  7. Predicciones_Total  Idem con el modelo aplicado a TODA la data
                         (in-sample, sesgado/optimista — sanity check).
  8. Glosario            Diccionario de terminos tecnicos en lenguaje claro.

Formato condicional en columna "Error porcentual (%)":
  rojo (>20%), amarillo (10-20%), verde (<10%).
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

from src.config import (
    ARTIFACTS_DIR,
    REPORT_BUSINESS_UNIT,
    REPORT_MAE_AMBER_RATIO,
    REPORT_MAE_TARGET,
    REPORT_PROJECT_NAME,
    REPORT_R2_AMBER_THRESHOLD,
    REPORT_R2_TARGET,
)
from src.step_05_evaluate.explainability import (
    build_context,
    compute_verdict,
    glossary_terms,
    kpi_explanatory_power,
    kpi_precision,
    kpi_vs_baseline,
    recommended_actions,
)


# ---------------------------------------------------------------------------
# Renombrado de columnas (modelo/error -> lenguaje claro)
# ---------------------------------------------------------------------------

# Mapeo aplicado a las hojas Predicciones_*. Solo renombramos columnas
# tecnicas del modelo; las del dominio (FORMATO, FUNDO, FECHA, KG/HA, ...)
# se quedan como estan porque el negocio ya las conoce con esos nombres.
_PRETTY_COLUMNS: Dict[str, str] = {
    "row_id": "ID fila",
    "KG/JR_H_real": "Productividad real (kg/jornal-hora)",
    "KG/JR_H_pred": "Productividad estimada (kg/jornal-hora)",
    "H-EF": "Horas efectivas (jornada)",
    "KG/JR_real": "Cosecha real (kg/jornal)",
    "KG/JR_estimado": "Cosecha estimada (kg/jornal)",
    "error_abs (kg/jornal)": "Error absoluto (kg/jornal)",
    "error_pct (%)": "Error porcentual (%)",
    "fold_id": "Partición CV",
}


# ---------------------------------------------------------------------------
# Construccion de DataFrames
# ---------------------------------------------------------------------------


def _build_predictions_df(
    X_raw: pd.DataFrame,
    business_cols: pd.DataFrame,
    y_h_true: np.ndarray,
    y_h_pred: np.ndarray,
    fold_id: Optional[np.ndarray] = None,
) -> pd.DataFrame:
    """Une features + target + predicciones + error en KG/JR.

    Filas con KG/JR_H_pred = NaN (folds no cubiertos en OOF) o con
    H-EF / KG/JR NaN se descartan: no se pueden traducir a unidad de negocio.
    """
    h_ef = business_cols["H-EF"].to_numpy(dtype=float)
    kg_jr_real = business_cols["KG/JR"].to_numpy(dtype=float)
    y_h_pred = np.asarray(y_h_pred, dtype=float)
    y_h_true = np.asarray(y_h_true, dtype=float)

    kg_jr_pred = y_h_pred * h_ef
    abs_err = np.abs(kg_jr_pred - kg_jr_real)
    with np.errstate(divide="ignore", invalid="ignore"):
        pct_err = np.where(kg_jr_real != 0, abs_err / np.abs(kg_jr_real) * 100.0, np.nan)

    df = X_raw.reset_index(drop=True).copy()
    df.insert(0, "row_id", np.arange(len(df)))
    df["KG/JR_H_real"] = y_h_true
    df["KG/JR_H_pred"] = y_h_pred
    df["H-EF"] = h_ef
    df["KG/JR_real"] = kg_jr_real
    df["KG/JR_estimado"] = kg_jr_pred
    df["error_abs (kg/jornal)"] = abs_err
    df["error_pct (%)"] = pct_err
    if fold_id is not None:
        df["fold_id"] = np.asarray(fold_id, dtype=int)

    mask = (
        np.isfinite(y_h_pred)
        & np.isfinite(h_ef)
        & np.isfinite(kg_jr_real)
    )
    df = df.loc[mask].reset_index(drop=True)

    # Renombrado a lenguaje claro
    return df.rename(columns=_PRETTY_COLUMNS)


def _build_subgroup_summary(df: pd.DataFrame, group_col: str) -> Optional[pd.DataFrame]:
    """Agrega metricas por subgrupo y rankea de peor a mejor por error %."""
    if group_col not in df.columns:
        return None
    err_abs_col = "Error absoluto (kg/jornal)"
    err_pct_col = "Error porcentual (%)"
    real_col = "Cosecha real (kg/jornal)"
    pred_col = "Cosecha estimada (kg/jornal)"
    g = df.groupby(group_col, dropna=False)
    out = pd.DataFrame({
        "n cosechas": g.size(),
        "Real promedio (kg/jornal)": g[real_col].mean(),
        "Estimado promedio (kg/jornal)": g[pred_col].mean(),
        "MAE (kg/jornal)": g[err_abs_col].mean(),
        "Error % promedio": g[err_pct_col].mean(),
        "Error máximo (kg/jornal)": g[err_abs_col].max(),
    })
    out = out.sort_values("Error % promedio", ascending=False).round(3)
    out.insert(0, "Ranking (peor → mejor)", np.arange(1, len(out) + 1))
    return out


def _build_resumen_df(
    *,
    variety: str,
    model_type: str,
    nested_metrics: Dict[str, float],
    business_validation,
    timestamp: str,
    n_oof: int,
    n_insample: int,
) -> pd.DataFrame:
    """Resumen tecnico para DS / auditoria. Lenguaje tecnico OK aqui."""
    moof = business_validation.metrics_oof if business_validation else {}
    mins = business_validation.metrics_insample if business_validation else {}
    r2_kgjr = float(moof.get("r2", float("nan")))
    mae_kgh = float(nested_metrics.get("nested_cv_mae_mean", float("nan")))

    def _state_r2(v):
        if v != v: return "—"
        return (
            "VERDE" if v >= REPORT_R2_TARGET
            else ("AMARILLO" if v >= REPORT_R2_AMBER_THRESHOLD else "ROJO")
        )

    def _state_mae(v):
        if v != v: return "—"
        return (
            "VERDE" if v <= REPORT_MAE_TARGET
            else (
                "AMARILLO" if v <= REPORT_MAE_AMBER_RATIO * REPORT_MAE_TARGET
                else "ROJO"
            )
        )

    rows = [
        ("## METADATA ##", "", ""),
        ("Variedad", variety, ""),
        ("Modelo", model_type.upper(), ""),
        ("Generado", timestamp, ""),
        ("Filas evaluadas (OOF)", n_oof, ""),
        ("Filas evaluadas (Total)", n_insample, ""),
        ("", "", ""),
        ("## TARGETS GERENCIALES ##", "", ""),
        ("Target R² (KG/JR, OOF) >=", f"{REPORT_R2_TARGET:.2f}", ""),
        ("Target MAE (KG/JR_H, Test CV) <=", f"{REPORT_MAE_TARGET:.2f}", ""),
        ("", "", ""),
        ("## METRICAS DEL MODELO (KG/JR_H) ##", "", ""),
        ("MAE Test CV", f"{nested_metrics.get('nested_cv_mae_mean', float('nan')):.4f}", _state_mae(mae_kgh)),
        ("MAE std Test CV", f"{nested_metrics.get('nested_cv_mae_std', float('nan')):.4f}", ""),
        ("MAE Train CV", f"{nested_metrics.get('nested_cv_mae_train_mean', float('nan')):.4f}", ""),
        ("Brecha (Test - Train)", f"{nested_metrics.get('nested_cv_gap_mean', float('nan')):+.4f}", ""),
        ("R² Test CV", f"{nested_metrics.get('nested_cv_r2_mean', float('nan')):.4f}", ""),
        ("", "", ""),
        ("## METRICAS EN UNIDAD DE NEGOCIO (KG/JR) ##", "", ""),
        ("R² OOF (gerencial)", f"{moof.get('r2', float('nan')):.4f}", _state_r2(r2_kgjr)),
        ("MAE OOF (kg/jornal)", f"{moof.get('mae', float('nan')):.4f}", ""),
        ("RMSE OOF", f"{moof.get('rmse', float('nan')):.4f}", ""),
        ("MAPE OOF (%)", f"{moof.get('mape', float('nan')):.2f}", ""),
        ("", "", ""),
        ("R² Aplicación Total (sanity)", f"{mins.get('r2', float('nan')):.4f}", ""),
        ("MAE Aplicación Total", f"{mins.get('mae', float('nan')):.4f}", ""),
        ("MAPE Aplicación Total (%)", f"{mins.get('mape', float('nan')):.2f}", ""),
    ]
    return pd.DataFrame(rows, columns=["Métrica", "Valor", "Estado"])


def _build_inicio_df(
    *,
    variety: str,
    model_type: str,
    timestamp: str,
    business_validation,
    abs_gap: float,
    n_rows: int,
) -> pd.DataFrame:
    """Hoja portada en lenguaje natural. Lo unico que un gerente necesita leer."""
    moof = business_validation.metrics_oof if business_validation else {}
    real_oof = getattr(business_validation, "kg_jr_real_oof", None)
    pred_oof = getattr(business_validation, "kg_jr_pred_oof", None)
    abs_err = (
        np.abs(np.asarray(pred_oof) - np.asarray(real_oof))
        if real_oof is not None and pred_oof is not None and len(real_oof) > 0
        else np.array([])
    )
    # Veredicto + KPIs en lenguaje natural usan SIEMPRE metricas OOF
    # (honestas, sobre datos no vistos). In-sample seria optimista.
    oof_mape = float(moof.get("mape", float("nan")))
    oof_r2 = moof.get("r2")

    verdict = compute_verdict(full_mape_pct=oof_mape, abs_gap=abs_gap)
    k1 = kpi_precision(abs_err, oof_mape)
    k2 = kpi_explanatory_power(oof_r2)
    k3 = kpi_vs_baseline(real_oof if real_oof is not None else np.array([]),
                         pred_oof if pred_oof is not None else np.array([]))

    rows = [
        ("◆◆◆ DASHBOARD EJECUTIVO ◆◆◆", "", ""),
        (REPORT_PROJECT_NAME, "", ""),
        ("", "", ""),
        ("Variedad", variety, ""),
        ("Modelo elegido", model_type.upper(), ""),
        ("Unidad de negocio", REPORT_BUSINESS_UNIT, ""),
        ("Cosechas analizadas", f"{n_rows:,}", ""),
        ("Generado", timestamp, ""),
        ("", "", ""),
        ("◆◆◆ VEREDICTO ◆◆◆", "", ""),
        (f"{verdict.icon}  {verdict.title}", "", verdict.color_key.upper()),
        (verdict.headline, "", ""),
        (verdict.body, "", ""),
        ("", "", ""),
        ("◆◆◆ LAS 3 PREGUNTAS QUE IMPORTAN ◆◆◆", "", ""),
        ("", "", ""),
        ("1. ¿Qué tan preciso es?", "", k1.score_label),
        (k1.headline, "", ""),
        (k1.detail, "", ""),
        ("", "", ""),
        ("2. ¿Cuánto explica?", "", k2.score_label),
        (k2.headline, "", ""),
        (k2.detail, "", ""),
        ("", "", ""),
        ("3. ¿Vale la pena vs no usar modelo?", "", k3.score_label),
        (k3.headline, "", ""),
        (k3.detail, "", ""),
        ("", "", ""),
        ("◆◆◆ ÍNDICE DE HOJAS ◆◆◆", "", ""),
        ("Inicio", "Esta hoja: veredicto + KPIs en lenguaje simple", ""),
        ("Acciones", "Qué hacer hoy con esta información", ""),
        ("Resumen", "Métricas técnicas globales (auditoría / DS)", ""),
        ("Por_FORMATO", "Error promedio por formato del producto", ""),
        ("Por_FUNDO", "Error promedio por fundo", ""),
        ("Predicciones_OOF", "Detalle fila a fila — predicciones honestas (no contaminadas)", ""),
        ("Predicciones_Total", "Detalle fila a fila — modelo aplicado a TODA la data (sanity)", ""),
        ("Glosario", "Términos técnicos explicados en lenguaje simple", ""),
    ]
    return pd.DataFrame(rows, columns=["Concepto", "Valor", "Nota"])


def _build_acciones_df(
    *,
    business_validation,
    abs_gap: float,
    full_mape: float,
    X_aligned: Optional[pd.DataFrame],
) -> pd.DataFrame:
    """Tabla de acciones recomendadas auto-generada."""
    real_oof = getattr(business_validation, "kg_jr_real_oof", None)
    pred_oof = getattr(business_validation, "kg_jr_pred_oof", None)
    real = np.asarray(real_oof if real_oof is not None else [])
    pred = np.asarray(pred_oof if pred_oof is not None else [])
    abs_errs = np.abs(pred - real) if real.size > 0 else np.array([])
    global_mape = float(business_validation.metrics_oof.get("mape", float("nan"))) if business_validation else float("nan")

    actions = recommended_actions(
        abs_errors=abs_errs, real=real, X_aligned=X_aligned,
        global_mape=global_mape, abs_gap=abs_gap, full_mape=full_mape,
    )
    rows = []
    for i, a in enumerate(actions, 1):
        sev_label = {"critical": "CRÍTICO", "warning": "ATENCIÓN", "info": "OK"}.get(a.severity, a.severity.upper())
        rows.append((f"#{i}", f"{a.icon} {sev_label}", a.title, a.body))
    if not rows:
        rows = [("—", "OK", "Sin acciones recomendadas", "")]
    return pd.DataFrame(rows, columns=["#", "Severidad", "Acción", "Detalle"])


def _build_glosario_df() -> pd.DataFrame:
    """Hoja de glosario para lectores no-tecnicos."""
    rows = [(t, d) for t, d in glossary_terms()]
    return pd.DataFrame(rows, columns=["Término", "Significado"])


# ---------------------------------------------------------------------------
# Formato del Excel (openpyxl)
# ---------------------------------------------------------------------------


_VERDICT_COLORS = {
    "GREEN":   ("C8E6C9", "1B5E20"),  # bg, fg
    "GREEN-2": ("DCEDC8", "33691E"),
    "AMBER":   ("FFE0B2", "BF360C"),
    "RED":     ("FFCDD2", "B71C1C"),
}

_SCORE_COLORS = {
    "ALTO":  ("C8E6C9", "1B5E20"),
    "MEDIO": ("FFE0B2", "BF360C"),
    "BAJO":  ("FFCDD2", "B71C1C"),
}

_SEVERITY_COLORS = {
    "CRÍTICO": ("FFCDD2", "B71C1C"),
    "ATENCIÓN": ("FFE0B2", "BF360C"),
    "OK": ("C8E6C9", "1B5E20"),
}


def _apply_formatting(workbook) -> None:
    """Aplica estilos consistentes a todas las hojas del Excel."""
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="0C2A4D")  # navy
    header_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill("solid", fgColor="EEF3FA")
    section_font = Font(bold=True, color="0C2A4D", size=12)
    block_fill = PatternFill("solid", fgColor="C9A961")  # gold
    block_font = Font(bold=True, color="0C2A4D", size=14)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if ws.max_row < 1:
            continue

        # Header (fila 1) en navy + bold blanco
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"

        # Auto-width (cap 60 para hojas con texto largo, 40 para tablas)
        cap = 80 if sheet_name in ("Inicio", "Acciones", "Glosario") else 45
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
                default=10,
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(cap, max(12, max_len + 2))

        # Wrap text en las hojas narrativas
        if sheet_name in ("Inicio", "Acciones", "Glosario"):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                ws.row_dimensions[row[0].row].height = None  # auto

        # Hojas de predicciones: formato condicional en error %
        if sheet_name.startswith("Predicciones_"):
            header = [c.value for c in ws[1]]
            err_col_name = "Error porcentual (%)"
            if err_col_name in header:
                col_idx = header.index(err_col_name) + 1
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator="greaterThan", formula=["20"],
                    fill=PatternFill("solid", fgColor="FCE4E4"),
                    font=Font(color="9B1C1C", bold=True)))
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator="between", formula=["10", "20"],
                    fill=PatternFill("solid", fgColor="FFF5D6")))
                ws.conditional_formatting.add(rng, CellIsRule(
                    operator="lessThan", formula=["10"],
                    fill=PatternFill("solid", fgColor="E0F2EE")))

        # Hoja Resumen: pintar secciones ##  ##
        if sheet_name == "Resumen":
            for row in ws.iter_rows(min_row=2):
                first = row[0].value
                if isinstance(first, str) and first.startswith("##"):
                    for c in row:
                        c.fill = section_fill
                        c.font = section_font

        # Hoja Inicio: pintar bloques ◆◆◆ y celdas de veredicto/score
        if sheet_name == "Inicio":
            for row in ws.iter_rows(min_row=2):
                first = row[0].value
                third = row[2].value if len(row) >= 3 else None
                if isinstance(first, str) and first.startswith("◆◆◆"):
                    for c in row[:3]:
                        c.fill = block_fill
                        c.font = block_font
                        c.alignment = Alignment(horizontal="left", vertical="center")
                    ws.row_dimensions[row[0].row].height = 28
                # Pintar celda de veredicto/score
                if isinstance(third, str):
                    key = third.upper()
                    if key in _VERDICT_COLORS:
                        bg, fg = _VERDICT_COLORS[key]
                        for c in row[:3]:
                            c.fill = PatternFill("solid", fgColor=bg)
                            c.font = Font(bold=True, color=fg, size=12)
                    elif key in _SCORE_COLORS:
                        bg, fg = _SCORE_COLORS[key]
                        row[2].fill = PatternFill("solid", fgColor=bg)
                        row[2].font = Font(bold=True, color=fg, size=11)
                        row[2].alignment = Alignment(horizontal="center")

        # Hoja Acciones: pintar columna severidad
        if sheet_name == "Acciones":
            header = [c.value for c in ws[1]]
            if "Severidad" in header:
                sev_col = header.index("Severidad") + 1
                for row in ws.iter_rows(min_row=2):
                    cell = row[sev_col - 1]
                    val = (cell.value or "").upper()
                    for key, (bg, fg) in _SEVERITY_COLORS.items():
                        if key in val:
                            cell.fill = PatternFill("solid", fgColor=bg)
                            cell.font = Font(bold=True, color=fg)
                            cell.alignment = Alignment(horizontal="center")
                            break

        # Borde sutil en todas las celdas con datos (excepto bloques pintados)
        if sheet_name not in ("Inicio",):
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value not in (None, ""):
                        cell.border = thin_border


# ---------------------------------------------------------------------------
# API publica
# ---------------------------------------------------------------------------


def export_business_excel(
    *,
    variety: str,
    model_type: str,
    X_raw: pd.DataFrame,
    business_cols: pd.DataFrame,
    oof: Dict[str, np.ndarray],
    final_pipeline,
    business_validation,
    nested_metrics: Dict[str, float],
    output_dir: Path | str | None = None,
    filename: Optional[str] = None,
) -> Optional[Path]:
    """Genera el Excel ejecutivo multi-hoja del campeon.

    Devuelve la ruta del archivo o None si faltan KG/JR / H-EF.
    """
    if (
        business_cols is None
        or "H-EF" not in business_cols.columns
        or "KG/JR" not in business_cols.columns
    ):
        return None

    out_dir = Path(output_dir) if output_dir else ARTIFACTS_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = filename or f"business_export_{variety}_{model_type}_{ts}.xlsx"
    out_path = out_dir / fname

    # ---- DataFrames ----
    df_oof = _build_predictions_df(
        X_raw=X_raw, business_cols=business_cols,
        y_h_true=oof["y_true"], y_h_pred=oof["y_pred"],
        fold_id=oof.get("fold_id"),
    )
    try:
        y_h_pred_full = final_pipeline.predict(X_raw)
    except Exception:
        logger.warning(
            "business_export: final_pipeline.predict(X_raw) fallo; "
            "Predicciones_Total saldra con NaN", exc_info=True,
        )
        y_h_pred_full = np.full(len(X_raw), np.nan)
    df_total = _build_predictions_df(
        X_raw=X_raw, business_cols=business_cols,
        y_h_true=np.asarray(oof["y_true"], dtype=float),
        y_h_pred=y_h_pred_full,
    )

    by_formato = _build_subgroup_summary(df_oof, "FORMATO")
    by_fundo = _build_subgroup_summary(df_oof, "FUNDO")

    df_resumen = _build_resumen_df(
        variety=variety, model_type=model_type,
        nested_metrics=nested_metrics,
        business_validation=business_validation,
        timestamp=datetime.now().isoformat(timespec="seconds"),
        n_oof=len(df_oof), n_insample=len(df_total),
    )

    # Veredicto + acciones usan OOF (honesto). abs_gap viene de nested CV.
    abs_gap = abs(float(nested_metrics.get("nested_cv_gap_mean", 0.0)))
    oof_metrics = (business_validation.metrics_oof or {}) if business_validation else {}
    oof_mape = float(oof_metrics.get("mape", float("nan")))

    # X alineado con OOF (para subgroups en Acciones)
    X_aligned: Optional[pd.DataFrame] = None
    bv_mask = getattr(business_validation, "oof_mask", None)
    if bv_mask is not None:
        try:
            X_aligned = X_raw.iloc[np.asarray(bv_mask, dtype=bool)].reset_index(drop=True)
        except Exception:
            X_aligned = None

    df_inicio = _build_inicio_df(
        variety=variety, model_type=model_type,
        timestamp=datetime.now().strftime("%Y-%m-%d %H:%M"),
        business_validation=business_validation,
        abs_gap=abs_gap,
        n_rows=len(df_oof),
    )
    df_acciones = _build_acciones_df(
        business_validation=business_validation,
        abs_gap=abs_gap, full_mape=oof_mape,
        X_aligned=X_aligned,
    )
    df_glosario = _build_glosario_df()

    # ---- Escritura ordenada (de simple a detallado) ----
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_inicio.to_excel(writer, sheet_name="Inicio", index=False)
        df_acciones.to_excel(writer, sheet_name="Acciones", index=False)
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        if by_formato is not None:
            by_formato.to_excel(writer, sheet_name="Por_FORMATO")
        if by_fundo is not None:
            by_fundo.to_excel(writer, sheet_name="Por_FUNDO")
        df_oof.to_excel(writer, sheet_name="Predicciones_OOF", index=False)
        df_total.to_excel(writer, sheet_name="Predicciones_Total", index=False)
        df_glosario.to_excel(writer, sheet_name="Glosario", index=False)
        _apply_formatting(writer.book)

    return out_path
